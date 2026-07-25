import os

import openpyxl
from openpyxl.utils import get_column_letter

from config import CONFIG

EXCEL_HEADER = [
    "公司名", "税号", "发票号码", "重复标记",
    "发票类型", "金额", "开票日期", "新文件名", "原文件名", "车牌号",
]


def _char_width(s):
    """计算字符串显示宽度（中文算2，英文算1）"""
    return sum(2 if ord(c) > 127 else 1 for c in str(s))


def write_excel_batch(path, excel_rows_by_sheet):
    """批量写入所有发票数据到 Excel，只打开/保存一次文件"""
    if not excel_rows_by_sheet:
        return
    excel_path = os.path.join(path, CONFIG["excel_filename"])
    wb = openpyxl.load_workbook(excel_path) if os.path.exists(excel_path) else openpyxl.Workbook()

    for sheet_name, rows in excel_rows_by_sheet.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(EXCEL_HEADER)

        existing_rows = set()
        for existing in ws.iter_rows():
            values = tuple(cell.value if cell.value is not None else "" for cell in existing)
            existing_rows.add(values)

        for row in rows:
            row_tuple = tuple(row)
            if row_tuple in existing_rows:
                continue
            ws.append(row)
            existing_rows.add(row_tuple)

        for col_idx in range(len(EXCEL_HEADER)):
            letter = get_column_letter(col_idx + 1)
            max_width = _char_width(EXCEL_HEADER[col_idx])
            for row in rows:
                max_width = max(max_width, _char_width(row[col_idx]))
            ws.column_dimensions[letter].width = max_width + 3

    wb.save(excel_path)
    wb.close()
